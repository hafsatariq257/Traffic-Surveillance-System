import openpyxl
import os
from datetime import datetime
 
EXCEL_FILE = "violations.xlsx"
 
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Violations"
        ws.append(["#", "Timestamp", "Violation Type", "Number Plate", "Vehicle Class", "Speed (km/h)"])
        wb.save(EXCEL_FILE)
 
def log_violation(violation_type, plate_text, vehicle_class, speed):
    init_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    row_num = ws.max_row
    ws.append([
        row_num,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        violation_type,
        plate_text if plate_text else "Unknown",
        vehicle_class,
        round(speed, 2)
    ])
    wb.save(EXCEL_FILE)
 
def get_violations():
    init_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            data.append(row)
    return data
