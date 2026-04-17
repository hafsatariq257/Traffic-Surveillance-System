"""
Traffic Monitoring System — Core Detection Engine
Handles: vehicle detection, speed, counting, signal violation,
         helmet detection, number plate OCR, Excel logging.
"""

import cv2
import numpy as np
import time
import re
import math
from pathlib import Path
from datetime import datetime
from collections import defaultdict, deque
from dataclasses import dataclass, field
from typing import Optional

import easyocr
import openpyxl
from ultralytics import YOLO

# ─── DATA CLASSES ─────────────────────────────────────────────────────────────

@dataclass
class TrackedVehicle:
    track_id:     int
    class_name:   str
    bbox:         tuple               # (x1,y1,x2,y2)
    centroid:     tuple               # (cx,cy)
    speed_kmh:    float = 0.0
    has_helmet:   bool  = True        # default: assume yes
    plate_text:   str   = ""
    violated:     bool  = False
    last_seen:    float = field(default_factory=time.time)
    positions:    deque = field(default_factory=lambda: deque(maxlen=30))
    timestamps:   deque = field(default_factory=lambda: deque(maxlen=30))


@dataclass
class ViolationRecord:
    timestamp:   str
    track_id:    int
    vehicle_cls: str
    plate_text:  str
    speed_kmh:   float
    violation:   str


# ─── SPEED ESTIMATOR ──────────────────────────────────────────────────────────

class SpeedEstimator:
    """
    Estimates speed using pixel displacement between frames.
    Requires calibration: measure a known real-world distance in pixels.
    """

    def __init__(self, fps: float, pixels_per_meter: float = 8.0):
        """
        fps              : video frames per second
        pixels_per_meter : calibration — how many pixels = 1 metre in the scene.
                           Measure a road marking (e.g. 3 m lane width) in pixels.
        """
        self.fps              = fps
        self.pixels_per_meter = pixels_per_meter

    def update(self, vehicle: TrackedVehicle) -> float:
        """Add current centroid and return smoothed speed in km/h."""
        now = time.time()
        vehicle.positions.append(vehicle.centroid)
        vehicle.timestamps.append(now)

        if len(vehicle.positions) < 5:
            return 0.0

        # Use last 10 points for smoothed speed
        n = min(10, len(vehicle.positions))
        p1 = vehicle.positions[-n]
        p2 = vehicle.positions[-1]
        t1 = vehicle.timestamps[-n]
        t2 = vehicle.timestamps[-1]

        dt = t2 - t1
        if dt < 1e-6:
            return vehicle.speed_kmh

        pixel_dist = math.hypot(p2[0] - p1[0], p2[1] - p1[1])
        meters     = pixel_dist / self.pixels_per_meter
        speed_ms   = meters / dt
        speed_kmh  = speed_ms * 3.6

        # Clamp unrealistic values
        speed_kmh = min(max(speed_kmh, 0.0), 200.0)

        # Exponential smoothing
        alpha = 0.3
        vehicle.speed_kmh = alpha * speed_kmh + (1 - alpha) * vehicle.speed_kmh
        return vehicle.speed_kmh


# ─── SIGNAL VIOLATION DETECTOR ────────────────────────────────────────────────

class SignalViolationDetector:
    """
    Detects if a vehicle crosses a stop-line when signal is RED.
    Stop-line is defined as a horizontal y-coordinate in the frame.
    """

    def __init__(self, stop_line_y: int, signal_red: bool = True):
        self.stop_line_y  = stop_line_y
        self.signal_red   = signal_red
        self._crossed     = set()   # track IDs that already crossed

    def set_signal(self, red: bool):
        self.signal_red = red
        if not red:
            self._crossed.clear()   # reset when light turns green

    def check(self, vehicle: TrackedVehicle) -> bool:
        """Returns True if this vehicle just committed a violation."""
        if not self.signal_red:
            return False
        if vehicle.track_id in self._crossed:
            return False

        cy = vehicle.centroid[1]
        if cy > self.stop_line_y:
            self._crossed.add(vehicle.track_id)
            vehicle.violated = True
            return True
        return False


# ─── NUMBER PLATE DETECTOR + OCR ──────────────────────────────────────────────

class NumberPlateReader:
    """
    1. Detects plate region using a secondary YOLO model (or heuristic crop).
    2. Runs EasyOCR on the cropped plate.
    """

    # Common Pakistani / generic plate pattern
    PLATE_PATTERN = re.compile(r"[A-Z]{2,3}[-\s]?\d{3,4}", re.IGNORECASE)

    def __init__(self, plate_model_path: Optional[str] = None, languages=("en",)):
        self.reader = easyocr.Reader(list(languages), gpu=False, verbose=False)
        self.plate_model = YOLO(plate_model_path) if plate_model_path else None

    def _crop_plate(self, frame: np.ndarray, bbox: tuple) -> Optional[np.ndarray]:
        """Crop plate region: use plate model if available, else bottom-centre heuristic."""
        x1, y1, x2, y2 = [int(v) for v in bbox]
        h = y2 - y1
        w = x2 - x1

        if self.plate_model:
            crop = frame[y1:y2, x1:x2]
            results = self.plate_model.predict(crop, imgsz=320, conf=0.3, verbose=False)
            if results and results[0].boxes:
                box = results[0].boxes[0].xyxy[0].cpu().numpy().astype(int)
                return crop[box[1]:box[3], box[0]:box[2]]

        # Heuristic: bottom-centre 60% width, bottom 30% height
        px1 = x1 + int(w * 0.2)
        px2 = x2 - int(w * 0.2)
        py1 = y1 + int(h * 0.70)
        py2 = y2
        plate = frame[py1:py2, px1:px2]
        return plate if plate.size > 0 else None

    def read(self, frame: np.ndarray, bbox: tuple) -> str:
        plate_img = self._crop_plate(frame, bbox)
        if plate_img is None or plate_img.size == 0:
            return ""

        # Pre-process for better OCR
        gray = cv2.cvtColor(plate_img, cv2.COLOR_BGR2GRAY)
        gray = cv2.resize(gray, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
        _, thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

        results = self.reader.readtext(thresh, allowlist="ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-")
        if not results:
            return ""

        # Pick highest-confidence result
        best = max(results, key=lambda r: r[2])
        text = best[1].upper().strip()

        # Validate against plate pattern (optional — comment out if too strict)
        # if not self.PLATE_PATTERN.search(text):
        #     return text   # return anyway, unvalidated

        return text


# ─── HELMET DETECTOR ──────────────────────────────────────────────────────────

class HelmetDetector:
    """
    Uses a secondary YOLO model trained to detect 'helmet' / 'no_helmet'.
    Falls back to colour-heuristic if no model provided.
    """

    def __init__(self, helmet_model_path: Optional[str] = None):
        self.model = YOLO(helmet_model_path) if helmet_model_path else None

    def detect(self, frame: np.ndarray, bbox: tuple, class_name: str) -> bool:
        """Returns True if helmet is detected (or vehicle is not a motorcycle)."""
        if class_name not in ("motorcycle", "heavy_bike"):
            return True   # not applicable

        x1, y1, x2, y2 = [int(v) for v in bbox]
        h = y2 - y1
        w = x2 - x1

        # Crop rider's head region (top 35% of bounding box)
        head_crop = frame[y1:y1 + int(h * 0.35), x1:x2]
        if head_crop.size == 0:
            return True

        if self.model:
            results = self.model.predict(head_crop, imgsz=320, conf=0.4, verbose=False)
            if results and results[0].boxes:
                names = self.model.names
                for box in results[0].boxes:
                    cls_id = int(box.cls[0])
                    if names[cls_id] == "helmet":
                        return True
                    if names[cls_id] == "no_helmet":
                        return False
            return True   # no detection → assume helmet

        # Colour heuristic fallback: dark blob at top of crop = helmet
        gray  = cv2.cvtColor(head_crop, cv2.COLOR_BGR2GRAY)
        dark  = np.sum(gray < 80)
        total = gray.size
        ratio = dark / (total + 1e-6)
        return ratio > 0.15   # >15% dark pixels ≈ helmet


# ─── EXCEL LOGGER ─────────────────────────────────────────────────────────────

class ViolationLogger:
    def __init__(self, output_path: str = "violations.xlsx"):
        self.path = Path(output_path)
        self.wb   = openpyxl.Workbook()
        self.ws   = self.wb.active
        self.ws.title = "Violations"

        # Header
        headers = ["Timestamp", "Track ID", "Vehicle Class",
                   "Number Plate", "Speed (km/h)", "Violation Type"]
        self.ws.append(headers)

        # Style header
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", fgColor="1F3864")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in self.ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        self._row_count = 0

    def log(self, record: ViolationRecord):
        self._row_count += 1
        row = [
            record.timestamp,
            record.track_id,
            record.vehicle_cls,
            record.plate_text or "Undetected",
            round(record.speed_kmh, 1),
            record.violation,
        ]
        self.ws.append(row)

        # Alternate row colour
        from openpyxl.styles import PatternFill
        if self._row_count % 2 == 0:
            fill = PatternFill("solid", fgColor="DCE6F1")
            for cell in self.ws[self.ws.max_row]:
                cell.fill = fill

        self.save()

    def save(self):
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(self.path)


# ─── VEHICLE COUNTER ──────────────────────────────────────────────────────────

class VehicleCounter:
    """Counts unique track IDs that cross a counting line."""

    def __init__(self, count_line_y: int):
        self.count_line_y = count_line_y
        self._counted     = set()
        self.counts       = defaultdict(int)
        self.total        = 0

    def update(self, vehicle: TrackedVehicle) -> bool:
        """Returns True if this vehicle was counted for the first time."""
        if vehicle.track_id in self._counted:
            return False
        cy = vehicle.centroid[1]
        if cy > self.count_line_y:
            self._counted.add(vehicle.track_id)
            self.counts[vehicle.class_name] += 1
            self.total += 1
            return True
        return False


# ─── ANNOTATOR (draws on frame) ───────────────────────────────────────────────

COLORS = {
    "car":        (0,   200, 255),
    "motorcycle": (255, 128,   0),
    "bicycle":    (0,   255, 128),
    "auto":       (255, 255,   0),
    "truck":      (128,   0, 255),
    "tanker":     (255,   0, 128),
    "heavy_bike": (0,   128, 255),
    "default":    (200, 200, 200),
}

def draw_vehicle(frame, vehicle: TrackedVehicle, signal_red: bool):
    x1, y1, x2, y2 = [int(v) for v in vehicle.bbox]
    color = COLORS.get(vehicle.class_name, COLORS["default"])

    # Red border for violators
    border_color = (0, 0, 255) if vehicle.violated else color
    thickness    = 3 if vehicle.violated else 2
    cv2.rectangle(frame, (x1, y1), (x2, y2), border_color, thickness)

    # Label background
    label = (f"{vehicle.class_name} #{vehicle.track_id} "
             f"{vehicle.speed_kmh:.0f}km/h")
    (tw, th), _ = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.5, 1)
    cv2.rectangle(frame, (x1, y1 - th - 8), (x1 + tw + 4, y1), border_color, -1)
    cv2.putText(frame, label, (x1 + 2, y1 - 4),
                cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)

    # Plate text below box
    if vehicle.plate_text:
        plate_label = f"PLATE: {vehicle.plate_text}"
        cv2.putText(frame, plate_label, (x1, y2 + 16),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 255), 1)

    # Helmet warning
    if not vehicle.has_helmet:
        cv2.putText(frame, "NO HELMET!", (x1, y1 - th - 22),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 2)

    # VIOLATION flash
    if vehicle.violated:
        cv2.putText(frame, "SIGNAL VIOLATION!", (x1, y2 + 32),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 255), 2)


def draw_ui(frame, counter: VehicleCounter, signal_red: bool,
            stop_line_y: int, count_line_y: int):
    h, w = frame.shape[:2]

    # Stop line
    line_color = (0, 0, 255) if signal_red else (0, 255, 0)
    cv2.line(frame, (0, stop_line_y), (w, stop_line_y), line_color, 2)
    cv2.putText(frame, "STOP LINE", (10, stop_line_y - 6),
                cv2.FONT_HERSHEY_SIMPLEX, 0.5, line_color, 1)

    # Count line
    cv2.line(frame, (0, count_line_y), (w, count_line_y), (0, 255, 255), 1)
    cv2.putText(frame, "COUNT LINE", (10, count_line_y - 6),
                cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 255), 1)

    # Signal indicator
    sig_text  = "SIGNAL: RED" if signal_red else "SIGNAL: GREEN"
    sig_color = (0, 0, 255) if signal_red else (0, 255, 0)
    cv2.rectangle(frame, (w - 200, 10), (w - 10, 50), (30, 30, 30), -1)
    cv2.putText(frame, sig_text, (w - 195, 36),
                cv2.FONT_HERSHEY_SIMPLEX, 0.7, sig_color, 2)

    # Vehicle counts panel
    panel_x = 10
    panel_y = 10
    cv2.rectangle(frame, (panel_x, panel_y), (panel_x + 220, panel_y + 30 + len(counter.counts) * 22 + 22),
                  (30, 30, 30), -1)
    cv2.putText(frame, f"TOTAL: {counter.total}", (panel_x + 8, panel_y + 24),
                cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)
    for i, (cls, cnt) in enumerate(counter.counts.items()):
        cv2.putText(frame, f"  {cls}: {cnt}", (panel_x + 8, panel_y + 46 + i * 22),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, COLORS.get(cls, (200,200,200)), 1)

    return frame